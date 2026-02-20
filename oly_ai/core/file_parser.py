# Copyright (c) 2026, OLY Technologies and contributors
# File Parser — Extract text from uploaded files (PDF, Excel, CSV, TXT, etc.)
# Used by the chat API to let AI analyze file attachments.

import csv
import io
import os

import frappe
from frappe import _


# Supported file extensions and their parsers
SUPPORTED_EXTENSIONS = {
	".pdf",
	".xlsx", ".xls",
	".csv",
	".txt", ".md", ".log", ".json", ".xml", ".html",
	".docx",
}

# Max text to extract per file (chars) — prevents huge docs from blowing up context
MAX_EXTRACT_CHARS = 50_000


def parse_file(file_url):
	"""Parse a Frappe file URL and extract its text content.

	Args:
		file_url: Frappe file URL like /files/report.pdf or /private/files/data.xlsx

	Returns:
		dict: {
			"filename": str,
			"extension": str,
			"text": str,         # Extracted text content
			"truncated": bool,   # Whether text was truncated
			"pages": int|None,   # Number of pages (PDF only)
			"rows": int|None,    # Number of rows (CSV/Excel only)
		}
	"""
	# Resolve file path on disk
	fpath = _resolve_file_path(file_url)
	if not fpath or not os.path.exists(fpath):
		return {"error": f"File not found: {file_url}"}

	filename = os.path.basename(fpath)
	ext = os.path.splitext(filename)[1].lower()

	if ext not in SUPPORTED_EXTENSIONS:
		return {"error": f"Unsupported file type: {ext}. Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"}

	try:
		if ext == ".pdf":
			return _parse_pdf(fpath, filename)
		elif ext in (".xlsx", ".xls"):
			return _parse_excel(fpath, filename, ext)
		elif ext == ".csv":
			return _parse_csv(fpath, filename)
		elif ext == ".docx":
			return _parse_docx(fpath, filename)
		else:
			# Plain text files
			return _parse_text(fpath, filename, ext)
	except Exception as e:
		frappe.logger("oly_ai").debug(f"File parse error ({filename}): {e}")
		return {"error": f"Failed to parse {filename}: {str(e)}"}


def _resolve_file_path(file_url):
	"""Convert a Frappe file URL to an absolute file path."""
	if not file_url:
		return None

	site_path = frappe.get_site_path()

	if file_url.startswith("/files/"):
		return os.path.join(site_path, "public", file_url.lstrip("/"))
	elif file_url.startswith("/private/files/"):
		return os.path.join(site_path, file_url.lstrip("/"))
	elif file_url.startswith("/api/method/"):
		# Not a direct file
		return None
	else:
		# Try as-is under public
		return os.path.join(site_path, "public", "files", os.path.basename(file_url))


def _truncate(text, filename):
	"""Truncate text if over limit, return (text, truncated)."""
	if len(text) > MAX_EXTRACT_CHARS:
		return text[:MAX_EXTRACT_CHARS] + f"\n\n... [Truncated — {filename} exceeds {MAX_EXTRACT_CHARS:,} char limit]", True
	return text, False


def _parse_pdf(fpath, filename):
	"""Extract text from a PDF file."""
	try:
		from PyPDF2 import PdfReader
	except ImportError:
		return {"error": "PyPDF2 not installed. Run: pip install PyPDF2"}

	reader = PdfReader(fpath)
	pages = len(reader.pages)
	text_parts = []

	for i, page in enumerate(reader.pages):
		page_text = page.extract_text() or ""
		if page_text.strip():
			text_parts.append(f"--- Page {i + 1} ---\n{page_text.strip()}")

	text = "\n\n".join(text_parts)
	if not text.strip():
		return {"error": f"Could not extract text from {filename}. The PDF may be image-based (scanned)."}

	text, truncated = _truncate(text, filename)
	return {
		"filename": filename,
		"extension": ".pdf",
		"text": text,
		"truncated": truncated,
		"pages": pages,
		"rows": None,
	}


def _parse_excel(fpath, filename, ext):
	"""Extract data from an Excel file as formatted text."""
	try:
		from openpyxl import load_workbook
	except ImportError:
		return {"error": "openpyxl not installed. Run: pip install openpyxl"}

	wb = load_workbook(fpath, read_only=True, data_only=True)
	text_parts = []
	total_rows = 0

	for sheet_name in wb.sheetnames:
		ws = wb[sheet_name]
		rows = list(ws.iter_rows(values_only=True))
		if not rows:
			continue

		total_rows += len(rows)
		text_parts.append(f"=== Sheet: {sheet_name} ({len(rows)} rows) ===")

		# Format as markdown table
		for i, row in enumerate(rows[:500]):  # Limit to 500 rows per sheet
			cells = [str(c) if c is not None else "" for c in row]
			line = " | ".join(cells)
			text_parts.append(line)
			if i == 0:
				# Add separator after header
				text_parts.append(" | ".join(["---"] * len(cells)))

		if len(rows) > 500:
			text_parts.append(f"... [{len(rows) - 500} more rows omitted]")

	wb.close()

	text = "\n".join(text_parts)
	text, truncated = _truncate(text, filename)
	return {
		"filename": filename,
		"extension": ext,
		"text": text,
		"truncated": truncated,
		"pages": None,
		"rows": total_rows,
	}


def _parse_csv(fpath, filename):
	"""Extract data from a CSV file as formatted text."""
	with open(fpath, "r", encoding="utf-8", errors="replace") as f:
		# Sniff dialect
		sample = f.read(8192)
		f.seek(0)
		try:
			dialect = csv.Sniffer().sniff(sample)
		except csv.Error:
			dialect = csv.excel

		reader = csv.reader(f, dialect)
		rows = []
		for i, row in enumerate(reader):
			if i >= 1000:  # Limit to 1000 rows
				break
			rows.append(row)

	if not rows:
		return {"error": f"No data found in {filename}"}

	total_rows = len(rows)
	text_parts = [f"=== CSV: {filename} ({total_rows} rows) ==="]

	for i, row in enumerate(rows):
		line = " | ".join(str(c) for c in row)
		text_parts.append(line)
		if i == 0:
			text_parts.append(" | ".join(["---"] * len(row)))

	text = "\n".join(text_parts)
	text, truncated = _truncate(text, filename)
	return {
		"filename": filename,
		"extension": ".csv",
		"text": text,
		"truncated": truncated,
		"pages": None,
		"rows": total_rows,
	}


def _parse_docx(fpath, filename):
	"""Extract text from a Word document (basic — paragraph text only)."""
	try:
		import zipfile
		import xml.etree.ElementTree as ET

		with zipfile.ZipFile(fpath) as z:
			with z.open("word/document.xml") as f:
				tree = ET.parse(f)

		ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
		paragraphs = []
		for p in tree.iter(f"{{{ns['w']}}}p"):
			texts = [t.text for t in p.iter(f"{{{ns['w']}}}t") if t.text]
			if texts:
				paragraphs.append("".join(texts))

		text = "\n\n".join(paragraphs)
		if not text.strip():
			return {"error": f"No text found in {filename}"}

		text, truncated = _truncate(text, filename)
		return {
			"filename": filename,
			"extension": ".docx",
			"text": text,
			"truncated": truncated,
			"pages": None,
			"rows": None,
		}
	except Exception as e:
		return {"error": f"Could not parse {filename}: {str(e)}"}


def _parse_text(fpath, filename, ext):
	"""Read plain text files."""
	with open(fpath, "r", encoding="utf-8", errors="replace") as f:
		text = f.read()

	text, truncated = _truncate(text, filename)
	return {
		"filename": filename,
		"extension": ext,
		"text": text,
		"truncated": truncated,
		"pages": None,
		"rows": None,
	}


def parse_files_for_context(file_urls):
	"""Parse multiple files and build a context string for AI injection.

	Args:
		file_urls: List of Frappe file URLs

	Returns:
		str: Formatted context text with all file contents, ready for system prompt injection
	"""
	if not file_urls:
		return ""

	parts = []
	for url in file_urls:
		result = parse_file(url)
		if result.get("error"):
			parts.append(f"📎 {url}: {result['error']}")
		else:
			header = f"📎 File: {result['filename']}"
			if result.get("pages"):
				header += f" ({result['pages']} pages)"
			elif result.get("rows"):
				header += f" ({result['rows']} rows)"
			if result.get("truncated"):
				header += " [truncated]"
			parts.append(f"{header}\n\n{result['text']}")

	return "\n\n" + "═" * 60 + "\n\n".join(parts)
